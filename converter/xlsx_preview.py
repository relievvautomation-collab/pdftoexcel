"""Read first-sheet preview rows from an .xlsx file for API/UI."""

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read_xlsx_preview(
    path: str | Path,
    *,
    max_rows: int = 120,
    max_cols: int = 52,
) -> Tuple[List[str], List[List[str]]]:
    """
    Return column letters as headers (A,B,…) and cell values as strings.
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [get_column_letter(i) for i in range(1, max_cols + 1)]
        rows_out: List[List[str]] = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_rows,
            max_col=max_cols,
            values_only=True,
        ):
            rows_out.append(["" if v is None else str(v) for v in row])
        return headers, rows_out
    finally:
        wb.close()
