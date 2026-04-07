"""Build .xlsx: extracted tables sheet + per-page PDF raster + optional layout sheet."""

from __future__ import annotations

import io
import math
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage

from .color_utils import rgb_to_hex
from .image_handler import image_to_png_bytes
from .pdf_parser import LineDrawing, ParsedPDF, TextBlock
from .pdf_visual_theme import (
    SheetTheme,
    apply_worksheet_page_setup,
    build_theme_for_page,
    default_sheet_theme,
)
from .statement_excel import (
    RowKind,
    build_amount_column_mask,
    classify_statement_rows,
    dedupe_tds_extraction_rows,
    try_parse_number,
)
from .traces_metadata import extract_assessee_header, extract_traces_preamble, is_traces_pdf
from .traces_normalize import find_status_of_booking_column, normalize_traces_table_columns


GRID_COLS = 78
MIN_ROWS_PAGE = 72
MAX_ROWS_PAGE = 360
POINT_SCALE = 0.75

PREVIEW_MAX_ROWS = 120
PREVIEW_MAX_COLS = 52


def _merge_anchor(ws, row: int, col: int) -> Tuple[int, int]:
    """Top-left cell for a coordinate (only that cell accepts .value in a merge)."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return m.min_row, m.min_col
    return row, col


# Raster quality knobs:
# - Increase RASTER_DPI for sharper text/colors (larger file, slower conversion).
# - Increase MAX_IMAGE_WIDTH_PX to reduce downscaling on very wide pages.
RASTER_DPI = 300
MAX_IMAGE_WIDTH_PX = 5000  # larger cap = less scaling = closer PDF match
PAGE_GAP_ROWS = 2


def _thin_border(color: str = "000000") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _side_from_line(w: float, dashed: bool) -> str:
    if dashed:
        return "dashed"
    if w >= 2.0:
        return "medium"
    return "thin"


def _build_border_from_drawings(
    drawings: List[LineDrawing],
    page: int,
    grid_top: int,
    pw: float,
    ph: float,
    rows_for_page: int,
    scale_x: float,
    scale_y: float,
) -> Dict[Tuple[int, int], Border]:
    out: Dict[Tuple[int, int], Border] = {}
    for ln in drawings:
        if ln.page != page:
            continue
        rel_r = min(rows_for_page - 1, max(0, int(ln.y0 / ph * rows_for_page)))
        rel_c = min(GRID_COLS - 1, max(0, int(ln.x0 / pw * GRID_COLS)))
        r = grid_top + rel_r
        c = rel_c + 1
        key = (min(max(1, r), 1048576), min(max(1, c), 16384))
        side_style = _side_from_line(ln.width, ln.dashed)
        col = rgb_to_hex(ln.color_rgb)
        side = Side(style=side_style, color=col)
        if abs(ln.y1 - ln.y0) < 1.0:
            out[key] = Border(bottom=side)
        elif abs(ln.x1 - ln.x0) < 1.0:
            out[key] = Border(right=side)
    return out


def _map_bbox_to_cells(
    pw: float,
    ph: float,
    rows_for_page: int,
    x0: float,
    y0: float,
    x1: float,
    y1: float,
    grid_top: int,
) -> Tuple[int, int, int, int]:
    c1 = min(GRID_COLS, max(1, int(x0 / max(pw, 1e-6) * GRID_COLS) + 1))
    c2 = min(GRID_COLS, max(c1, int(math.ceil(x1 / max(pw, 1e-6) * GRID_COLS))))
    r1_rel = min(rows_for_page - 1, max(0, int(y0 / max(ph, 1e-6) * rows_for_page)))
    r2_rel = min(rows_for_page - 1, max(r1_rel, int(y1 / max(ph, 1e-6) * rows_for_page)))
    if r2_rel < r1_rel:
        r2_rel = r1_rel
    r1 = grid_top + r1_rel
    r2 = grid_top + r2_rel
    return r1, c1, r2, c2


def _range_free(r1: int, c1: int, r2: int, c2: int, occupied: Set[Tuple[int, int]]) -> bool:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            if (r, c) in occupied:
                return False
    return True


def _occupy_range(r1: int, c1: int, r2: int, c2: int, occupied: Set[Tuple[int, int]]) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            occupied.add((r, c))


def _style_cell(
    cell,
    tb: TextBlock,
    border_map: Dict[Tuple[int, int], Border],
    r: int,
    c: int,
) -> None:
    hex_font = rgb_to_hex(tb.color_rgb)
    cell.font = Font(
        name=(tb.font.split("+")[-1][:31] if tb.font else "Calibri"),
        size=max(6, min(48, round(tb.size))),
        bold=tb.bold,
        italic=tb.italic,
        underline="single" if tb.underline else None,
        color=hex_font,
    )
    al = "left"
    if tb.align == "center":
        al = "center"
    elif tb.align == "right":
        al = "right"
    cell.alignment = Alignment(horizontal=al, vertical="top", wrap_text=True)
    if (r, c) in border_map:
        cell.border = border_map[(r, c)]


def _pixmap_to_png_bytes(page) -> bytes:
    """Render page at high DPI — preserves colours, images, layout like the PDF."""
    try:
        pix = page.get_pixmap(dpi=RASTER_DPI, alpha=False)
    except (TypeError, AttributeError):
        z = RASTER_DPI / 72.0
        pix = page.get_pixmap(matrix=fitz.Matrix(z, z), alpha=False)
    return pix.tobytes("png")


def _structured_table_nonempty(data: List[List[str]]) -> bool:
    return any(any(str(c).strip() for c in row) for row in (data or []))


def _structured_table_sig(page: int, data: List[List[str]]) -> str:
    if not data:
        return f"{page}|0|0|"
    nc = max(len(r) for r in data)
    head = tuple(tuple(str(c)[:80] for c in r[:10]) for r in data[:4])
    return f"{page}|{len(data)}|{nc}|{head}"


def _collect_structured_tables(parsed: ParsedPDF) -> List[Dict[str, Any]]:
    """Merge Camelot + pdfplumber tables; dedupe near-identical grids."""
    seen: Set[str] = set()
    out: List[Dict[str, Any]] = []

    for tab in parsed.camelot_tables:
        raw = tab.get("data") or []
        data = [[str(c) if c is not None else "" for c in row] for row in raw]
        if not _structured_table_nonempty(data):
            continue
        sig = _structured_table_sig(int(tab.get("page", 0)), data)
        if sig in seen:
            continue
        seen.add(sig)
        out.append(
            {
                "page": int(tab.get("page", 0)),
                "source": "camelot",
                "flavor": tab.get("flavor"),
                "data": data,
            }
        )

    for tab in parsed.plumber_tables:
        data = tab.get("data") or []
        if not _structured_table_nonempty(data):
            continue
        sig = _structured_table_sig(int(tab.get("page", 0)), data)
        if sig in seen:
            continue
        seen.add(sig)
        out.append(
            {
                "page": int(tab.get("page", 0)),
                "source": "pdfplumber",
                "data": data,
            }
        )

    out.sort(key=lambda t: (t["page"], t["source"]))
    return out


def _merge_traces_camelot_tables(
    structured: List[Dict[str, Any]], parsed: ParsedPDF
) -> List[Dict[str, Any]]:
    """
    Multi-page 26AS PDFs often yield one Camelot table per page. Concatenate them
    into a single Table 1 so row order matches reading the PDF top-to-bottom.
    """
    if not is_traces_pdf(parsed):
        return structured
    camelot = [t for t in structured if t.get("source") == "camelot"]
    rest = [t for t in structured if t.get("source") != "camelot"]
    if len(camelot) <= 1:
        return structured
    camelot.sort(key=lambda t: (int(t.get("page", 0)), str(t.get("flavor", ""))))
    all_rows: List[List[str]] = []
    for tab in camelot:
        for row in tab.get("data") or []:
            all_rows.append([str(c) if c is not None else "" for c in row])
    if not all_rows:
        return structured
    max_c = max(len(r) for r in all_rows)
    padded = [r + [""] * (max_c - len(r)) for r in all_rows]
    merged: Dict[str, Any] = {
        "page": int(camelot[0].get("page", 0)),
        "source": "camelot",
        "flavor": camelot[0].get("flavor"),
        "data": padded,
    }
    pages_merged = {int(t.get("page", 0)) for t in camelot}
    rest_extra = [t for t in rest if int(t.get("page", 0)) not in pages_merged]
    return [merged] + rest_extra


def _row_height_for_kind(kind: RowKind, theme: Optional[SheetTheme] = None) -> float:
    scale = theme.row_scale if theme else 1.0
    if kind in (RowKind.PRIMARY_HEADER, RowKind.FALLBACK_HEADER):
        return 26.0 * scale
    if kind == RowKind.SUB_HEADER:
        return 30.0 * scale
    if kind == RowKind.MASTER_SUMMARY:
        return 34.0 * scale
    if kind == RowKind.DATA:
        return 24.0 * scale
    if kind == RowKind.TITLE:
        return 22.0 * scale
    return 20.0 * scale


def _style_cell_for_row_kind(
    cell: Any,
    kind: RowKind,
    amount_col: bool,
    raw_val: str,
    data_row_index: int = 0,
    traces_layout: bool = False,
    theme: Optional[SheetTheme] = None,
) -> None:
    t = theme if theme is not None else default_sheet_theme()
    cell.border = _thin_border()
    num = try_parse_number(raw_val)

    if kind in (RowKind.PRIMARY_HEADER, RowKind.FALLBACK_HEADER):
        cell.font = Font(
            bold=True, color=t.header_on_primary_hex, size=int(round(t.header_pt))
        )
        cell.fill = PatternFill(fill_type="solid", fgColor=t.primary_fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return

    if kind == RowKind.TITLE:
        cell.font = Font(bold=True, size=int(round(t.header_pt)), color=t.header_on_primary_hex)
        cell.fill = PatternFill(fill_type="solid", fgColor=t.title_fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return

    if kind == RowKind.SUB_HEADER:
        cell.font = Font(bold=True, color=t.body_font_hex, size=int(round(t.header_pt)))
        cell.fill = PatternFill(fill_type="solid", fgColor=t.sub_fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return

    if kind == RowKind.MASTER_SUMMARY:
        cell.fill = PatternFill(fill_type="solid", fgColor=t.master_fill_hex)
        cell.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
        v_master = "top" if traces_layout else "center"
        if amount_col and num is not None:
            cell.value = num
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical=v_master, wrap_text=True)
        return

    # DATA
    cell.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
    v_data = "top" if traces_layout else "center"
    if data_row_index % 2 == 0:
        cell.fill = PatternFill(fill_type="solid", fgColor=t.data_zebra_hex)
    if amount_col and num is not None:
        cell.value = num
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="left", vertical=v_data, wrap_text=True)


def _write_traces_assessee_grid(
    ws: Any,
    start_row: int,
    nc: int,
    meta: Dict[str, Any],
    theme: Optional[SheetTheme] = None,
) -> int:
    """
    Reference PDF layout: row 1 = four label/value pairs (PAN, Status, FY, AY);
    row 2 = Name; row 3 = Address.
    """
    t = theme if theme is not None else default_sheet_theme()
    r = start_row
    pairs = [
        ("Permanent Account Number (PAN)", str(meta.get("pan") or "")),
        ("Current Status of PAN", str(meta.get("pan_status") or "")),
        ("Financial Year", str(meta.get("financial_year") or "")),
        ("Assessment Year", str(meta.get("assessment_year") or "")),
    ]
    widths = [nc // 4] * 4
    for i in range(nc % 4):
        widths[i] += 1
    col = 1
    for seg in range(4):
        w = widths[seg]
        seg_end = col + w - 1
        if w <= 0:
            continue
        if w == 1:
            try:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=seg_end)
            except Exception:
                pass
            vc = ws.cell(row=r, column=col, value=pairs[seg][1])
            vc.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
            vc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            vc.border = _thin_border()
        else:
            lab_w = max(1, w // 2)
            lab_end = col + lab_w - 1
            val_start = lab_end + 1
            if val_start > seg_end:
                lab_end = seg_end - 1
                val_start = seg_end
            try:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=lab_end)
            except Exception:
                pass
            lb = ws.cell(row=r, column=col, value=pairs[seg][0])
            lb.font = Font(bold=True, size=int(round(t.label_pt)), color=t.header_on_primary_hex)
            lb.fill = PatternFill(fill_type="solid", fgColor=t.primary_fill_hex)
            lb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            lb.border = _thin_border()
            try:
                ws.merge_cells(start_row=r, start_column=val_start, end_row=r, end_column=seg_end)
            except Exception:
                pass
            vc = ws.cell(row=r, column=val_start, value=pairs[seg][1])
            vc.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
            vc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            vc.border = _thin_border()
        col = seg_end + 1
    ws.row_dimensions[r].height = 28.0 * t.row_scale
    r += 1

    lbl_span = min(3, max(2, nc // 6))
    val_start = lbl_span + 1
    try:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=lbl_span)
        ws.merge_cells(start_row=r, start_column=val_start, end_row=r, end_column=nc)
    except Exception:
        pass
    lb = ws.cell(row=r, column=1, value="Name of Assessee")
    lb.font = Font(bold=True, size=int(round(t.header_pt)), color=t.header_on_primary_hex)
    lb.fill = PatternFill(fill_type="solid", fgColor=t.primary_fill_hex)
    lb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    lb.border = _thin_border()
    vc = ws.cell(row=r, column=val_start, value=str(meta.get("name") or ""))
    vc.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
    vc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    vc.border = _thin_border()
    ws.row_dimensions[r].height = 28.0 * t.row_scale
    r += 1

    try:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=lbl_span)
        ws.merge_cells(start_row=r, start_column=val_start, end_row=r, end_column=nc)
    except Exception:
        pass
    lb = ws.cell(row=r, column=1, value="Address of Assessee")
    lb.font = Font(bold=True, size=int(round(t.header_pt)), color=t.header_on_primary_hex)
    lb.fill = PatternFill(fill_type="solid", fgColor=t.primary_fill_hex)
    lb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    lb.border = _thin_border()
    vc = ws.cell(row=r, column=val_start, value=str(meta.get("address") or ""))
    vc.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
    vc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    vc.border = _thin_border()
    addr_len = len(str(meta.get("address") or ""))
    ws.row_dimensions[r].height = min(120.0, max(36.0, 18.0 + (addr_len // 60) * 14.0)) * t.row_scale
    r += 1

    r += 1
    return r


def _write_annual_statement_header_fallback_assessee(
    ws: Any,
    start_row: int,
    nc: int,
    meta: Dict[str, Any],
    theme: Optional[SheetTheme] = None,
) -> int:
    """Six-row two-column assessee block when nc < 8."""
    t = theme if theme is not None else default_sheet_theme()
    r = start_row
    fields = [
        ("Permanent Account Number (PAN)", meta.get("pan") or ""),
        ("Current Status of PAN", meta.get("pan_status") or ""),
        ("Financial Year", meta.get("financial_year") or ""),
        ("Assessment Year", meta.get("assessment_year") or ""),
        ("Name of Assessee", meta.get("name") or ""),
        ("Address of Assessee", meta.get("address") or ""),
    ]
    for label, val in fields:
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=nc)
        except Exception:
            pass
        lb = ws.cell(row=r, column=1, value=label)
        lb.font = Font(bold=True, size=int(round(t.header_pt)), color=t.header_on_primary_hex)
        lb.fill = PatternFill(fill_type="solid", fgColor=t.primary_fill_hex)
        lb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        lb.border = _thin_border()
        vc = ws.cell(row=r, column=3, value=str(val) if val is not None else "")
        vc.font = Font(size=int(round(t.body_pt)), color=t.body_font_hex)
        vc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc.border = _thin_border()
        ws.row_dimensions[r].height = 24 * t.row_scale
        r += 1
    r += 1
    return r


def _write_annual_statement_header(
    ws: Any,
    start_row: int,
    ncols: int,
    meta: Dict[str, Any],
    theme: Optional[SheetTheme] = None,
) -> int:
    """Title, optional data-updated line, assessee label/value rows (TRACES-style). Returns next free row."""
    th = theme if theme is not None else default_sheet_theme()
    r = start_row
    nc = max(1, min(40, max(ncols, 12)))
    try:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    except Exception:
        pass
    title_cell = ws.cell(row=r, column=1, value="Annual Tax Statement")
    title_cell.font = Font(bold=True, size=int(round(th.title_pt)), color=th.header_on_primary_hex)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=th.primary_fill_hex)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = _thin_border()
    ws.row_dimensions[r].height = 28 * th.row_scale
    r += 1

    du = str(meta.get("data_updated_till") or "").strip()
    if du:
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        except Exception:
            pass
        c = ws.cell(row=r, column=1, value=f"Data updated till {du}")
        c.font = Font(size=int(round(th.small_pt)), color=th.muted_font_hex)
        c.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        c.border = _thin_border()
        ws.row_dimensions[r].height = 18 * th.row_scale
        r += 1

    if nc >= 8:
        r = _write_traces_assessee_grid(ws, r, nc, meta, theme=th)
    else:
        r = _write_annual_statement_header_fallback_assessee(ws, r, nc, meta, theme=th)
    return r


def _write_part_one_preamble(
    ws: Any,
    start_row: int,
    ncols: int,
    preamble: Dict[str, str],
    theme: Optional[SheetTheme] = None,
) -> int:
    """Disclaimer, PART-I title, INR note — between assessee block and extracted table (TRACES)."""
    th = theme if theme is not None else default_sheet_theme()
    r = start_row
    nc = max(1, min(40, ncols))
    disc = (preamble.get("pan_disclaimer") or "").strip()
    if disc:
        try:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        except Exception:
            pass
        dc = ws.cell(row=r, column=1, value=disc)
        dc.font = Font(size=int(round(th.small_pt)), color=th.body_font_hex)
        dc.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        dc.border = _thin_border()
        est_lines = max(2, (len(disc) // 100) + 1)
        ws.row_dimensions[r].height = min(120.0, float(14 + est_lines * 13)) * th.row_scale
        r += 1
        r += 1

    part_i = (preamble.get("part_i_title") or "PART-I - Details of Tax Deducted at Source").strip()
    try:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    except Exception:
        pass
    pc = ws.cell(row=r, column=1, value=part_i)
    pc.font = Font(bold=True, size=int(round(th.header_pt)), color=th.header_on_primary_hex)
    pc.fill = PatternFill(fill_type="solid", fgColor=th.primary_fill_hex)
    pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    pc.border = _thin_border()
    ws.row_dimensions[r].height = 26.0 * th.row_scale
    r += 1

    inr = (preamble.get("inr_note") or "(All amount values are in INR)").strip()
    try:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    except Exception:
        pass
    ic = ws.cell(row=r, column=1, value=inr)
    ic.font = Font(size=int(round(th.small_pt)), color=th.muted_font_hex)
    ic.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    ic.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ic.border = _thin_border()
    ws.row_dimensions[r].height = 18.0 * th.row_scale
    r += 1

    r += 1
    return r


def _fill_one_table_on_sheet(
    ws: Any,
    tab: Dict[str, Any],
    _idx: int,
    parsed: Optional[ParsedPDF] = None,
    theme: Optional[SheetTheme] = None,
) -> None:
    """One extracted table: 26AS-style bands when row keywords match; else fallback header + data."""
    th = theme if theme is not None else default_sheet_theme()
    col_scale = max(0.75, min(1.35, th.body_pt / 11.0))
    ws.sheet_view.showGridLines = True
    row = 1
    is_traces = bool(parsed is not None and is_traces_pdf(parsed))

    if _idx == 0 and parsed is not None:
        row = _prepend_page0_header_images(ws, parsed, start_row=1, traces_mode=is_traces)

    page = int(tab.get("page", 0)) + 1
    src = str(tab.get("source", ""))
    data = tab.get("data") or []
    if not data:
        ws.cell(row=row, column=1, value="(empty table)")
        apply_worksheet_page_setup(ws, th)
        return

    str_data: List[List[str]] = [[str(c) if c is not None else "" for c in r] for r in data]
    str_data = dedupe_tds_extraction_rows(str_data)
    if is_traces:
        str_data = normalize_traces_table_columns(str_data)
    ncols = min(40, max(len(r) for r in str_data))

    if is_traces and _idx == 0 and parsed is not None:
        meta = extract_assessee_header(parsed)
        row = _write_annual_statement_header(ws, row, ncols, meta, theme=th)
        row = _write_part_one_preamble(ws, row, max(ncols, 12), extract_traces_preamble(parsed), theme=th)
    else:
        title = f"Page {page} · {src}"
        try:
            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=max(1, ncols),
            )
        except Exception:
            pass
        tcell = ws.cell(row=row, column=1, value=title)
        tcell.font = Font(bold=True, size=int(round(th.header_pt)), color=th.header_on_primary_hex)
        tcell.fill = PatternFill(fill_type="solid", fgColor=th.title_fill_hex)
        tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1

    kinds = classify_statement_rows(str_data)
    header_candidates = [str_data[i] for i, k in enumerate(kinds) if k in (RowKind.PRIMARY_HEADER, RowKind.SUB_HEADER, RowKind.FALLBACK_HEADER)]
    amount_mask = build_amount_column_mask(header_candidates, ncols)
    status_col = find_status_of_booking_column(str_data) if is_traces else None

    data_only_index = 0
    for i, row_data in enumerate(str_data):
        kind = kinds[i] if i < len(kinds) else RowKind.DATA
        for j in range(ncols):
            val = row_data[j] if j < len(row_data) else ""
            cc = j + 1
            cell = ws.cell(row=row, column=cc)
            cell.value = val
            _style_cell_for_row_kind(
                cell,
                kind,
                amount_mask[j] if j < len(amount_mask) else False,
                val,
                data_row_index=data_only_index if kind == RowKind.DATA else 0,
                traces_layout=is_traces,
                theme=th,
            )
            if (
                is_traces
                and status_col is not None
                and j == status_col
                and kind == RowKind.DATA
                and str(val).strip().upper() == "F"
            ):
                cell.font = Font(size=int(round(th.body_pt)), color=th.ok_green_hex)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if kind == RowKind.DATA:
            data_only_index += 1
        ws.row_dimensions[row].height = _row_height_for_kind(kind, th)
        row += 1

    if is_traces and _idx == 0:
        note = ws.cell(row=row, column=1, value=f"Extraction source: Page {page} · {src}")
        note.font = Font(size=int(round(th.small_pt)), color="808080", italic=True)
        row += 1

    for c in range(1, min(ncols + 1, 41)):
        letter = get_column_letter(c)
        if is_traces and ncols >= 9:
            # Wider name/section columns; tighter codes/dates; amounts on the right
            if c == 1:
                w = 10.0
            elif c == 2:
                w = 40.0
            elif c == 3:
                w = 14.0
            elif c == 4:
                w = 18.0
            elif c == 5:
                w = 14.0
            elif c == 6:
                w = 14.0
            elif c <= len(amount_mask) and amount_mask[c - 1]:
                w = 15.0
            else:
                w = 16.0
        else:
            if c == 1:
                w = 12.0
            elif c == 2:
                w = 52.0
            elif c == 3:
                w = 22.0
            elif c <= len(amount_mask) and amount_mask[c - 1]:
                w = 16.0
            else:
                w = 20.0
        w = w * col_scale
        if ws.column_dimensions[letter].width is None or ws.column_dimensions[letter].width < 10:
            ws.column_dimensions[letter].width = w

    apply_worksheet_page_setup(ws, th)


def _fill_structured_tables_workbook(
    wb: Workbook,
    tables: List[Dict[str, Any]],
    parsed: Optional[ParsedPDF] = None,
) -> List[str]:
    """
    Numbers-style: one sheet per table (Table 1, Table 2, …) when tables exist.
    """
    names: List[str] = []
    if not tables:
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="No bordered tables detected in this PDF.")
        ws.cell(row=2, column=1, value="Use hidden 'PDF Page' sheets after download, or try a text-based PDF.")
        names.append(ws.title)
        return names

    max_sheets = 24
    use_tables = tables[:max_sheets]

    for i, tab in enumerate(use_tables):
        if i == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = f"Table {i + 1}"
        names.append(ws.title)
        page_idx = int(tab.get("page", 0))
        if parsed is not None and parsed.page_widths:
            page_idx = max(0, min(page_idx, len(parsed.page_widths) - 1))
        sheet_theme = (
            build_theme_for_page(parsed, page_idx, is_traces=is_traces_pdf(parsed))
            if parsed is not None
            else default_sheet_theme()
        )
        _fill_one_table_on_sheet(ws, tab, i, parsed=parsed, theme=sheet_theme)

    if len(tables) > max_sheets:
        ws = wb.create_sheet()
        ws.title = "More tables"
        names.append(ws.title)
        r = 1
        ws.cell(row=r, column=1, value=f"(+{len(tables) - max_sheets} tables omitted — raise max_sheets in code if needed.)")
    return names


def _prepend_page0_header_images(
    ws: Any,
    parsed: ParsedPDF,
    start_row: int = 1,
    traces_mode: bool = False,
) -> int:
    """Embed first-page PDF images (logos) above table on Table 1; return next free row."""
    page_imgs = [im for im in parsed.images if getattr(im, "page_index", -1) == 0][:5]
    if not page_imgs:
        return start_row
    row = start_row
    if not traces_mode:
        try:
            lab = ws.cell(row=row, column=1, value="Header images (from PDF page 1)")
            lab.font = Font(bold=True, size=10, color="1E3C72")
            row += 1
        except Exception:
            pass
    max_thumb_w = 200 if traces_mode else 320
    px_to_pt = 72.0 / 96.0

    def _place_image(im: Any, anchor_row: int, anchor_col: int) -> int:
        """Place image at anchor; return number of rows spanned."""
        anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
        raw = image_to_png_bytes(im.data, im.ext)
        pil = PILImage.open(io.BytesIO(raw))
        if pil.mode == "RGBA":
            bg = PILImage.new("RGB", pil.size, (255, 255, 255))
            bg.paste(pil, mask=pil.split()[3])
            pil = bg
        elif pil.mode not in ("RGB", "L"):
            pil = pil.convert("RGB")
        w0, h0 = pil.size
        if w0 < 1 or h0 < 1:
            return 1
        tw = min(max_thumb_w, w0)
        th = int(h0 * (tw / max(1, w0)))
        xl = XLImage(pil)
        xl.width = tw
        xl.height = th
        total_pt = float(xl.height) * px_to_pt
        rows_needed = max(1, int(math.ceil(total_pt / 15.0)))
        rh = total_pt / max(1, rows_needed)
        if rh > 409.0:
            rows_needed = max(1, int(math.ceil(total_pt / 409.0)))
            rh = total_pt / max(1, rows_needed)
        for rr in range(anchor_row, min(anchor_row + rows_needed, 1048576)):
            prev = ws.row_dimensions[rr].height
            if prev is None or prev < rh:
                ws.row_dimensions[rr].height = float(rh)
        ws.add_image(xl, anchor)
        return rows_needed

    if traces_mode:
        anchor_cols = (1, 9, 17)
        base_row = row
        max_span = 1
        for idx, im in enumerate(page_imgs[:3]):
            col = anchor_cols[min(idx, len(anchor_cols) - 1)]
            try:
                span = _place_image(im, base_row, col)
                max_span = max(max_span, span)
            except Exception:
                continue
        return base_row + max_span + 1

    for im in page_imgs:
        try:
            span = _place_image(im, row, 1)
            row += max(1, span) + 1
        except Exception:
            continue
    return row + 1


def _spread_columns_for_page_width(ws, n_cols: int, image_width_px: int) -> None:
    """
    Set column widths so the total grid width (approx) equals the image width.

    Excel column width approximation (Calibri 11):
      pixels ~= col_width * 7 + 5
    So for n columns:
      image_width_px / n ~= col_width * 7 + 5
    """
    per_px = float(image_width_px) / max(1, int(n_cols))
    per_w = (per_px - 5.0) / 7.0
    per_w = max(0.2, min(60.0, per_w))
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = per_w


def _append_embedded_page_images(
    ws: Any,
    parsed: ParsedPDF,
    pno: int,
    start_row: int,
) -> None:
    """Below the page raster, add PDF embedded images (logos etc.) with original colors."""
    page_imgs = [im for im in parsed.images if getattr(im, "page_index", -1) == pno]
    if not page_imgs:
        return
    row = max(1, start_row)
    try:
        label = ws.cell(row=row, column=1, value="Embedded images (from PDF)")
        label.font = Font(bold=True, size=11, color="1E3C72")
        row += 1
    except Exception:
        pass
    max_thumb_w = 480
    px_to_pt = 72.0 / 96.0
    for im in page_imgs:
        try:
            raw = image_to_png_bytes(im.data, im.ext)
            pil = PILImage.open(io.BytesIO(raw))
            if pil.mode == "RGBA":
                bg = PILImage.new("RGB", pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[3])
                pil = bg
            elif pil.mode not in ("RGB", "L"):
                pil = pil.convert("RGB")
            w0, h0 = pil.size
            if w0 < 1 or h0 < 1:
                continue
            tw = min(max_thumb_w, w0)
            th = int(h0 * (tw / max(1, w0)))
            xl = XLImage(pil)
            xl.width = tw
            xl.height = th
            total_pt = float(xl.height) * px_to_pt
            rows_needed = max(1, int(math.ceil(total_pt / 15.0)))
            rh = total_pt / max(1, rows_needed)
            if rh > 409.0:
                rows_needed = max(1, int(math.ceil(total_pt / 409.0)))
                rh = total_pt / max(1, rows_needed)
            for rr in range(row, min(row + rows_needed, 1048576)):
                prev = ws.row_dimensions[rr].height
                if prev is None or prev < rh:
                    ws.row_dimensions[rr].height = float(rh)
            anchor = f"A{row}"
            ws.add_image(xl, anchor)
            row += rows_needed + 1
        except Exception:
            continue


def _fill_raster_scan_sheets(
    wb: Workbook,
    pdf_path: str,
    progress_cb: Optional[Callable[[int, str], None]],
    parsed: ParsedPDF,
) -> List[str]:
    """
    iLovePDF-like visual match:
    - Each PDF page becomes its own worksheet.
    - A high-DPI raster of the page is anchored at A1.
    - Gridlines off, zoom set, print setup fit-to-page.
    """
    doc = fitz.open(pdf_path)
    n_cols_layout = 52
    sheet_names: List[str] = []
    try:
        n = len(doc)
        for pno in range(n):
            if progress_cb:
                pct = 12 + int(35 * (pno + 1) / max(n, 1))
                progress_cb(pct, f"Rendering page {pno + 1}/{n} (PDF quality)")

            page = doc[pno]

            # Always append; first sheet is reserved for Table/Statement sheets.
            ws = wb.create_sheet(title=f"PDF Page {pno + 1}")
            sheet_names.append(ws.title)

            # PDF-viewer look
            ws.sheet_view.showGridLines = False
            ws.sheet_view.zoomScale = 100

            # Print setup: one page wide x one page tall, minimal margins
            try:
                ws.sheet_properties.pageSetUpPr.fitToPage = True
            except Exception:
                pass
            try:
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 1
            except Exception:
                pass
            try:
                ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.0, footer=0.0)
            except Exception:
                pass
            try:
                ws.print_options.gridLines = False
            except Exception:
                pass

            # Orientation inferred from PDF aspect
            try:
                ws.page_setup.orientation = "landscape" if float(page.rect.width) > float(page.rect.height) else "portrait"
            except Exception:
                pass

            png_bytes = _pixmap_to_png_bytes(page)
            pil = PILImage.open(io.BytesIO(png_bytes))
            if pil.mode not in ("RGB", "L"):
                pil = pil.convert("RGB")

            w0, h0 = pil.size
            target_width_px = int(min(MAX_IMAGE_WIDTH_PX, max(1, w0)))
            scale = float(target_width_px) / max(1, w0)

            xl_img = XLImage(pil)
            xl_img.width = int(target_width_px)
            xl_img.height = int(h0 * scale)

            _spread_columns_for_page_width(ws, n_cols_layout, xl_img.width)

            # Row heights: match total image height in points
            px_to_pt = 72.0 / 96.0
            total_pt = float(xl_img.height) * px_to_pt
            rows_needed = max(1, int(math.ceil(total_pt / 15.0)))
            row_h = total_pt / max(1, rows_needed)
            if row_h > 409.0:
                rows_needed = max(1, int(math.ceil(total_pt / 409.0)))
                row_h = total_pt / max(1, rows_needed)
            for rr in range(1, min(1 + rows_needed, 1048576)):
                ws.row_dimensions[rr].height = float(row_h)

            ws.add_image(xl_img, "A1")
            _append_embedded_page_images(ws, parsed, pno, start_row=rows_needed + 2)
    finally:
        doc.close()
    return sheet_names


def _fill_extracted_sheet(ws: Any, parsed: ParsedPDF) -> None:
    """Text / images / tables mapped to cells (editable layer)."""
    ws.title = "Editable text & tables"
    page_count = len(parsed.page_widths)
    row_cursor = 1

    for pno in range(page_count):
        pw = parsed.page_widths[pno] or 612.0
        ph = parsed.page_heights[pno] or 792.0
        rows_for_page = int(GRID_COLS * ph / max(pw, 1.0))
        rows_for_page = max(MIN_ROWS_PAGE, min(MAX_ROWS_PAGE, rows_for_page))

        ws.cell(row=row_cursor, column=1, value=f"--- Page {pno + 1} ---")
        try:
            ws.merge_cells(
                start_row=row_cursor,
                start_column=1,
                end_row=row_cursor,
                end_column=min(24, GRID_COLS),
            )
        except Exception:
            pass
        ws.cell(row=row_cursor, column=1).font = Font(bold=True, size=11, color="064A9C")
        row_cursor += 1
        grid_top = row_cursor

        border_map = _build_border_from_drawings(
            parsed.drawings, pno, grid_top, pw, ph, rows_for_page, GRID_COLS / pw, rows_for_page / ph
        )

        occupied: Set[Tuple[int, int]] = set()
        page_blocks = [b for b in parsed.text_blocks if b.page == pno]
        page_blocks.sort(key=lambda b: (b.y0, b.x0))

        for tb in page_blocks:
            r1, c1, r2, c2 = _map_bbox_to_cells(pw, ph, rows_for_page, tb.x0, tb.y0, tb.x1, tb.y1, grid_top)
            placed = False
            if r2 >= r1 and c2 >= c1 and _range_free(r1, c1, r2, c2, occupied):
                try:
                    if r2 > r1 or c2 > c1:
                        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                    _occupy_range(r1, c1, r2, c2, occupied)
                    top = ws.cell(row=r1, column=c1, value=tb.text)
                    hex_font = rgb_to_hex(tb.color_rgb)
                    top.font = Font(
                        name=(tb.font.split("+")[-1][:31] if tb.font else "Calibri"),
                        size=max(6, min(48, round(tb.size))),
                        bold=tb.bold,
                        italic=tb.italic,
                        underline="single" if tb.underline else None,
                        color=hex_font,
                    )
                    al = "left"
                    if tb.align == "center":
                        al = "center"
                    elif tb.align == "right":
                        al = "right"
                    top.alignment = Alignment(horizontal=al, vertical="top", wrap_text=True)
                    if (r1, c1) in border_map:
                        top.border = border_map[(r1, c1)]
                    placed = True
                except Exception:
                    placed = False
            if not placed:
                r, c = r1, c1
                ar, ac = _merge_anchor(ws, r, c)
                # Only the merge anchor accepts .value; other cells are read-only MergedCells
                if (ar, ac) in occupied:
                    cur = ws.cell(row=ar, column=ac).value
                    sep = "\n" if cur else ""
                    cell = ws.cell(row=ar, column=ac, value=f"{cur or ''}{sep}{tb.text}")
                    _style_cell(cell, tb, border_map, ar, ac)
                else:
                    _occupy_range(ar, ac, ar, ac, occupied)
                    cell = ws.cell(row=ar, column=ac, value=tb.text)
                    _style_cell(cell, tb, border_map, ar, ac)

        for im in parsed.images:
            if im.page_index != pno:
                continue
            ix0, iy0, ix1, iy1 = im.bbox
            r1, c1, r2, c2 = _map_bbox_to_cells(pw, ph, rows_for_page, ix0, iy0, ix1, iy1, grid_top)
            if not _range_free(r1, c1, r2, c2, occupied):
                r1, c1, r2, c2 = r1, c1, r1, c1
            try:
                raw = image_to_png_bytes(im.data, im.ext)
                pil = PILImage.open(io.BytesIO(raw))
                xl_img = XLImage(pil)
                w_px = max(24, im.width_px)
                h_px = max(24, im.height_px)
                xl_img.width = min(w_px, 900)
                xl_img.height = min(h_px, 900)
                if r2 > r1 or c2 > c1:
                    try:
                        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                    except Exception:
                        pass
                _occupy_range(r1, c1, r2, c2, occupied)
                anchor = f"{get_column_letter(min(c1, 16384))}{r1}"
                ws.add_image(xl_img, anchor)
            except Exception:
                continue

        # Camelot tables are written on "Extracted tables" sheet only (avoid duplicate).

        # Extracted-sheet layout depends heavily on correct Excel sizing.
        # Previously this used hard clamps + an arbitrary scale factor, which
        # can visually shift text relative to its PDF coordinates.
        #
        # Excel column widths are in "character units". A practical approximation:
        #   col_width ~= (pixels - 5) / 7
        # where pixels ~= points * dpi / 72.
        dpi = 96.0
        col_points = pw / float(GRID_COLS)
        col_px = col_points * dpi / 72.0
        col_w = (col_px - 5.0) / 7.0
        col_w = max(0.2, min(30.0, float(col_w)))

        for c in range(1, min(GRID_COLS + 1, 200)):
            letter = get_column_letter(c)
            ws.column_dimensions[letter].width = col_w

        # Row heights are already in points in openpyxl, so use the normalized
        # page-height division directly (no hard min/max clamps).
        row_h = ph / float(rows_for_page)
        row_h = max(2.0, min(60.0, float(row_h)))
        for r in range(grid_top, min(grid_top + rows_for_page + 5, 1048576)):
            ws.row_dimensions[r].height = row_h

        row_cursor = max(row_cursor, grid_top + rows_for_page + 3)


def _sheet_preview_values(ws: Any, max_r: int, max_c: int) -> Tuple[List[str], List[List[str]]]:
    headers = [get_column_letter(i) for i in range(1, max_c + 1)]
    rows_out: List[List[str]] = []
    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v))
        if any(x.strip() for x in row_vals):
            rows_out.append(row_vals)
    return headers, rows_out


def build_excel(
    parsed: ParsedPDF,
    out_path: str,
    pdf_path: str,
    progress_cb: Optional[Callable[[int, str], None]] = None,
) -> Dict[str, Any]:
    """
    Primary: Table 1…N — spreadsheet-style data (navy headers, grid) from PDF tables.
    Backup: PDF Page sheets (full-page scan) — hidden when tables were extracted.
    Hidden: Layout (approx).
    """
    page_count = len(parsed.page_widths)

    wb = Workbook()
    structured = _collect_structured_tables(parsed)
    structured = _merge_traces_camelot_tables(structured, parsed)

    if progress_cb:
        progress_cb(8, "Building spreadsheet tables")

    table_sheet_names = _fill_structured_tables_workbook(wb, structured, parsed)

    if progress_cb:
        progress_cb(18, "Rendering PDF pages (scan backup)")

    raster_sheet_names = _fill_raster_scan_sheets(wb, pdf_path, progress_cb, parsed)
    raster_ws_list = [ws for ws in wb.worksheets if (ws.title or "").startswith("PDF Page")]

    # When we have extracted tables, user expects a Numbers-like workbook — hide picture sheets.
    if len(structured) > 0:
        for ws in raster_ws_list:
            ws.sheet_state = "hidden"

    if progress_cb:
        progress_cb(55, "Building layout (approx) sheet")

    ws_ext = wb.create_sheet()
    _fill_extracted_sheet(ws_ext, parsed)
    ws_ext.title = "Layout (approx)"
    ws_ext.sheet_state = "hidden"

    try:
        wb.active = 0
    except Exception:
        pass

    if progress_cb:
        progress_cb(85, "Saving workbook")

    wb.save(out_path)

    ws_preview = wb.worksheets[0]
    max_r = min(ws_preview.max_row or 1, PREVIEW_MAX_ROWS)
    max_c = min(ws_preview.max_column or PREVIEW_MAX_COLS, PREVIEW_MAX_COLS)
    if max_c < 1:
        max_c = PREVIEW_MAX_COLS
    preview_headers, preview_rows = _sheet_preview_values(ws_preview, max_r, max_c)

    if progress_cb:
        progress_cb(100, "Done")

    all_names = table_sheet_names + raster_sheet_names + ["Layout (approx)"]
    return {
        "rows_written": sum(ws.max_row or 0 for ws in wb.worksheets),
        "cols_written": max((ws.max_column or 0) for ws in wb.worksheets),
        "preview_headers": preview_headers,
        "preview_rows": preview_rows[:300],
        "page_count": page_count,
        "used_ocr": parsed.used_ocr,
        "structured_table_count": len(structured),
        "sheets": all_names,
        "visual_match_note": "Open Table 1… for spreadsheet-style data (like Numbers). Full-page PDF scans are on hidden PDF Page sheets if you need pixel-perfect reference. Unhide in Excel: right-click sheet → Unhide.",
    }
